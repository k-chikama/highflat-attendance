from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify, session
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import json
from dateutil.relativedelta import relativedelta
from math import floor
import requests
import sys

# Firestore関連のインポート
from firestore_config import firestore_manager
from auth_firestore import firestore_auth_manager, firestore_login_required
from attendance_firestore import firestore_attendance_manager

# 下位互換性のために従来のインポートも保持
from firebase_config import firebase_db
from auth import auth_manager, login_required

# jpholidayのインポートを安全に行う
try:
    import jpholiday
    JPHOLIDAY_AVAILABLE = True
except (ImportError, TypeError) as e:
    print(f"jpholidayインポートエラー: {e}")
    JPHOLIDAY_AVAILABLE = False

app = Flask(__name__)
# セッション用の秘密鍵（環境変数から取得、なければランダム生成）
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# データファイルのパス（フォールバック用）
DATA_FILE = 'attendance_data.json'

# Vercel環境でのデータ永続化のための設定（フォールバック用）
GIST_ID = os.environ.get('GIST_ID')
GITHUB_TOKEN = os.environ.get('GITHUB_TOKEN')
USE_GIST = bool(GIST_ID and GITHUB_TOKEN)

# Firestoreを使用するかどうかの設定
USE_FIRESTORE = os.environ.get('USE_FIRESTORE', 'true').lower() == 'true'

def get_auth_manager():
    """適切な認証マネージャーを取得"""
    if USE_FIRESTORE and firestore_manager.is_available():
        return firestore_auth_manager
    else:
        return auth_manager

def get_attendance_manager():
    """適切な勤怠マネージャーを取得"""
    if USE_FIRESTORE and firestore_manager.is_available():
        return firestore_attendance_manager
    else:
        # 従来の形式をエミュレート
        return None

def get_login_required_decorator():
    """適切なログイン必須デコレータを取得"""
    if USE_FIRESTORE and firestore_manager.is_available():
        return firestore_login_required
    else:
        return login_required

# 動的にデコレータを設定
login_required_decorator = get_login_required_decorator()

def load_user_data(username: str):
    """特定ユーザーの勤怠データを読み込む"""
    attendance_mgr = get_attendance_manager()
    
    if attendance_mgr:
        # Firestore版
        return attendance_mgr.get_user_attendance_data(username)
    else:
        # 従来版
        all_data = load_data()
        return all_data.get('users', {}).get(username, {})

def save_user_data(username: str, user_data: dict):
    """特定ユーザーの勤怠データを保存する"""
    attendance_mgr = get_attendance_manager()
    
    if attendance_mgr:
        # Firestore版 - 各日付のデータを個別に保存
        for date_str, daily_data in user_data.items():
            for field, value in daily_data.items():
                attendance_mgr.update_user_attendance_data(username, date_str, field, value)
    else:
        # 従来版
        all_data = load_data()
        if 'users' not in all_data:
            all_data['users'] = {}
        if username not in all_data['users']:
            all_data['users'][username] = {}
        
        all_data['users'][username] = user_data
        save_data(all_data)

def load_data():
    """全体勤怠データを読み込む（従来版互換）"""
    attendance_mgr = get_attendance_manager()
    
    if attendance_mgr:
        # Firestore版では個別ユーザーデータの取得のみ対応
        print("DEBUG: Firestore版を使用中、個別ユーザーデータを使用してください")
        return {}
    
    # 従来版のロジック
    print(f"DEBUG: Firebase利用可能 = {firebase_db.is_available()}")
    print(f"DEBUG: GIST_ID = {GIST_ID}")
    print(f"DEBUG: GITHUB_TOKEN = {'設定済み' if GITHUB_TOKEN else '未設定'}")
    print(f"DEBUG: USE_GIST = {USE_GIST}")
    
    # Firebase優先で試行
    if firebase_db.is_available():
        data = firebase_db.load_data()
        if data:  # Firebaseから正常にデータを取得できた場合
            return data
    
    # Firebaseが失敗した場合、Gistを試行
    if USE_GIST:
        return load_data_from_gist()
    
    # 最後にローカルファイル（開発環境のみ）
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    return {}

def save_data(data):
    """全体勤怠データを保存する（従来版互換）"""
    attendance_mgr = get_attendance_manager()
    
    if attendance_mgr:
        # Firestore版では自動保存されるためここでは何もしない
        print("DEBUG: Firestore版使用中、自動保存済み")
        return
    
    # 従来版のロジック
    saved = False
    
    # Firebase優先で保存
    if firebase_db.is_available():
        if firebase_db.save_data(data):
            saved = True
            print("DEBUG: Firebaseに保存成功")
    
    # Firebaseが失敗した場合、Gistに保存
    if not saved and USE_GIST:
        save_data_to_gist(data)
        saved = True
        print("DEBUG: Gistに保存成功")
    
    # 開発環境ではローカルファイルにも保存
    if not saved:
        try:
            with open(DATA_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print("DEBUG: ローカルファイルに保存成功")
        except Exception as e:
            print(f"ERROR: ローカル保存失敗 - {str(e)}")

def load_data_from_gist():
    """GitHub Gistからデータを読み込む"""
    try:
        headers = {
            'Authorization': f'token {GITHUB_TOKEN}',
            'Accept': 'application/vnd.github.v3+json'
        }
        response = requests.get(f'https://api.github.com/gists/{GIST_ID}', headers=headers)
        if response.status_code == 200:
            gist_data = response.json()
            files = gist_data.get('files', {})
            for filename, file_data in files.items():
                if filename == 'attendance_data.json':
                    content = file_data.get('content', '{}')
                    return json.loads(content)
        return {}
    except Exception as e:
        print(f"Gist読み込みエラー: {e}")
        return {}

def save_data_to_gist(data):
    """GitHub Gistにデータを保存する"""
    try:
        print(f"DEBUG: Gist保存開始 - データサイズ: {len(json.dumps(data))}")
        headers = {
            'Authorization': f'token {GITHUB_TOKEN}',
            'Accept': 'application/vnd.github.v3+json'
        }
        
        # 現在のGistを取得
        response = requests.get(f'https://api.github.com/gists/{GIST_ID}', headers=headers)
        print(f"DEBUG: Gist取得レスポンス: {response.status_code}")
        if response.status_code != 200:
            print(f"Gist取得エラー: {response.status_code}")
            print(f"DEBUG: レスポンス内容: {response.text}")
            return
        
        gist_data = response.json()
        files = gist_data.get('files', {})
        
        # ファイル内容を更新
        files['attendance_data.json'] = {
            'content': json.dumps(data, ensure_ascii=False, indent=2)
        }
        
        # Gistを更新
        update_data = {
            'files': files
        }
        
        response = requests.patch(
            f'https://api.github.com/gists/{GIST_ID}',
            headers=headers,
            json=update_data
        )
        
        print(f"DEBUG: Gist更新レスポンス: {response.status_code}")
        if response.status_code != 200:
            print(f"Gist更新エラー: {response.status_code}")
            print(f"DEBUG: レスポンス内容: {response.text}")
        else:
            print("DEBUG: Gist更新成功")
            
    except Exception as e:
        print(f"Gist保存エラー: {e}")
        import traceback
        print(f"DEBUG: エラー詳細: {traceback.format_exc()}")

def check_holiday(date):
    """祝日チェック（jpholidayライブラリを使用）"""
    if not JPHOLIDAY_AVAILABLE:
        return False
    
    try:
        return jpholiday.is_holiday(date)
    except Exception as e:
        print(f"祝日チェックエラー: {e}")
        return False

def get_month_range(year, month):
    """指定された年月の開始日と終了日を取得"""
    start_date = datetime(year, month, 1).date()
    if month == 12:
        end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
    else:
        end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
    return start_date, end_date

def create_excel_report(year, month, data, user_display_name):
    """Excelレポートを作成（画像フォーマットに準拠）"""
    import calendar
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "template"
    
    # 和暦変換関数
    def to_wareki(y, m):
        if y > 2019 or (y == 2019 and m >= 5):
            return f"令和 {y-2018}"
        else:
            return f"平成 {y-1988 if y >= 1989 else y}"
    
    # スタイル定義
    title_font = Font(name='MS Gothic', size=16, bold=True)
    header_font = Font(name='MS Gothic', size=10, bold=True)
    normal_font = Font(name='MS Gothic', size=9)
    red_font = Font(name='MS Gothic', size=9, color="FF0000", bold=True)
    bold_font = Font(name='MS Gothic', size=9, bold=True)
    
    # 色定義
    table_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    header_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
    company_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # 罫線定義
    thin = Side(style='thin', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # 配置定義
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    top = Alignment(horizontal='left', vertical='top')
    
    # 列幅設定
    col_widths = [5, 10, 12, 14, 14, 14, 10, 14, 14, 18, 4, 10, 12, 14, 14, 14, 10, 14, 14, 18, 4]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w
    
    # 行高設定
    for r in range(1, 60):
        ws.row_dimensions[r].height = 22
    ws.row_dimensions[2].height = 32
    
    # タイトル
    ws.merge_cells('B2:T2')
    ws['B2'] = "作業時間報告書"
    ws['B2'].font = title_font
    ws['B2'].alignment = center
    
    # 年月度
    ws['B4'] = to_wareki(year, month)
    ws['C4'] = f"{year}"
    ws['D4'] = "年"
    ws['E4'] = f"{month}"
    ws['F4'] = "月度"
    for cell in ['B4', 'C4', 'D4', 'E4', 'F4']:
        ws[cell].font = header_font
        ws[cell].alignment = center
    
    # 会社情報
    ws['B7'] = "対応客先名"
    ws['C7'] = "株式会社LINE"
    ws['F7'] = "会社名"
    ws['G7'] = "HIGHFLAT"
    ws['I7'] = "氏名"
    ws['J7'] = user_display_name or ""
    
    # 会社情報のスタイル
    for label_cell in ['B7', 'F7', 'I7']:
        ws[label_cell].font = header_font
        ws[label_cell].alignment = center
    for value_cell in ['C7', 'G7', 'J7']:
        ws[value_cell].font = normal_font
        ws[value_cell].alignment = left
        ws[value_cell].fill = company_fill
    
    # テーブルヘッダー（左側と右側）
    headers = ["日付", "曜日", "出勤時間", "退勤時間", "実働時間", "交通費", "出発駅", "目的駅", "備考"]
    
    # 左側ヘッダー（B列～J列）
    for i, h in enumerate(headers):
        cell = ws.cell(row=13, column=2+i, value=h)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = thin_border
    
    # 右側ヘッダー（L列～T列）
    for i, h in enumerate(headers):
        cell = ws.cell(row=13, column=12+i, value=h)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = thin_border
    
    # 月の日付範囲を取得
    start_date, end_date = get_month_range(year, month)
    days = []
    current_date = start_date
    while current_date <= end_date:
        days.append(current_date)
        current_date += timedelta(days=1)
    
    # 左側（1-16日）と右側（17-31日）に分割
    left_days = days[:16]
    right_days = days[16:]
    
    total_work = 0
    total_travel = 0
    
    # 左側データ（B列～J列）
    for idx, d in enumerate(left_days):
        r = 14 + idx
        date_str = d.strftime('%Y-%m-%d')
        weekday = d.strftime('%a')
        jp_week = {'Mon':'月','Tue':'火','Wed':'水','Thu':'木','Fri':'金','Sat':'土','Sun':'日'}[weekday]
        
        # データ取得（Firestoreデータフォーマットに対応）
        att = data.get(date_str, {})
        check_in = att.get('check_in', '')
        check_out = att.get('check_out', '')
        break_time = float(att.get('break_time', '1.0') or 1.0)
        notes = att.get('notes', '')
        travel_cost = att.get('travel_cost', '')
        travel_from = att.get('travel_from', '')
        travel_to = att.get('travel_to', '')
        
        is_holiday = check_holiday(d)
        
        # 実働時間計算
        def get_minutes(t):
            if not t: return None
            try:
                h, m = map(int, t.split(':'))
                return h*60 + m
            except:
                return None
        
        work_min = None
        if check_in and check_out:
            in_min = get_minutes(check_in)
            out_min = get_minutes(check_out)
            if in_min is not None and out_min is not None:
                work_min = out_min - in_min - int(break_time*60)
                if work_min < 0: work_min = None
        
        if work_min:
            total_work += work_min
        
        try:
            if travel_cost:
                total_travel += float(travel_cost)
        except:
            pass
        
        # セルに値を設定
        ws.cell(row=r, column=2, value=d.day).font = normal_font
        
        # 曜日（土日祝は赤色）
        weekday_cell = ws.cell(row=r, column=3, value=jp_week)
        if jp_week in ['土','日'] or is_holiday:
            weekday_cell.font = red_font
        else:
            weekday_cell.font = normal_font
        
        ws.cell(row=r, column=4, value=check_in).font = normal_font
        ws.cell(row=r, column=5, value=check_out).font = normal_font
        ws.cell(row=r, column=6, value=(f"{work_min/60:.2f}" if work_min is not None else "")).font = normal_font
        ws.cell(row=r, column=7, value=travel_cost).font = normal_font
        ws.cell(row=r, column=8, value=travel_from).font = normal_font
        ws.cell(row=r, column=9, value=travel_to).font = normal_font
        ws.cell(row=r, column=10, value=notes).font = normal_font
        
        # スタイル適用
        for i in range(9):
            c = ws.cell(row=r, column=2+i)
            c.fill = table_fill
            c.border = thin_border
            c.alignment = center if i not in [8] else left  # 備考欄は左寄せ
    
    # 右側データ（L列～T列）
    for idx, d in enumerate(right_days):
        r = 14 + idx
        date_str = d.strftime('%Y-%m-%d')
        weekday = d.strftime('%a')
        jp_week = {'Mon':'月','Tue':'火','Wed':'水','Thu':'木','Fri':'金','Sat':'土','Sun':'日'}[weekday]
        
        # データ取得
        att = data.get(date_str, {})
        check_in = att.get('check_in', '')
        check_out = att.get('check_out', '')
        break_time = float(att.get('break_time', '1.0') or 1.0)
        notes = att.get('notes', '')
        travel_cost = att.get('travel_cost', '')
        travel_from = att.get('travel_from', '')
        travel_to = att.get('travel_to', '')
        
        is_holiday = check_holiday(d)
        
        # 実働時間計算
        work_min = None
        if check_in and check_out:
            in_min = get_minutes(check_in)
            out_min = get_minutes(check_out)
            if in_min is not None and out_min is not None:
                work_min = out_min - in_min - int(break_time*60)
                if work_min < 0: work_min = None
        
        if work_min:
            total_work += work_min
        
        try:
            if travel_cost:
                total_travel += float(travel_cost)
        except:
            pass
        
        # セルに値を設定
        ws.cell(row=r, column=12, value=d.day).font = normal_font
        
        # 曜日（土日祝は赤色）
        weekday_cell = ws.cell(row=r, column=13, value=jp_week)
        if jp_week in ['土','日'] or is_holiday:
            weekday_cell.font = red_font
        else:
            weekday_cell.font = normal_font
        
        ws.cell(row=r, column=14, value=check_in).font = normal_font
        ws.cell(row=r, column=15, value=check_out).font = normal_font
        ws.cell(row=r, column=16, value=(f"{work_min/60:.2f}" if work_min is not None else "")).font = normal_font
        ws.cell(row=r, column=17, value=travel_cost).font = normal_font
        ws.cell(row=r, column=18, value=travel_from).font = normal_font
        ws.cell(row=r, column=19, value=travel_to).font = normal_font
        ws.cell(row=r, column=20, value=notes).font = normal_font
        
        # スタイル適用
        for i in range(9):
            c = ws.cell(row=r, column=12+i)
            c.fill = table_fill
            c.border = thin_border
            c.alignment = center if i not in [8] else left  # 備考欄は左寄せ
    
    # 合計行（右側のみ）
    sum_row = 14+max(len(left_days), len(right_days))
    
    # 実働時間合計
    ws.cell(row=sum_row, column=15, value="計").font = bold_font
    ws.cell(row=sum_row, column=15).fill = header_fill
    ws.cell(row=sum_row, column=15).border = thin_border
    ws.cell(row=sum_row, column=15).alignment = center
    
    ws.cell(row=sum_row, column=16, value=f"{total_work/60:.2f}").font = bold_font
    ws.cell(row=sum_row, column=16).fill = header_fill
    ws.cell(row=sum_row, column=16).border = thin_border
    ws.cell(row=sum_row, column=16).alignment = center
    
    # 交通費合計
    ws.cell(row=sum_row, column=18, value="交通費合計").font = bold_font
    ws.cell(row=sum_row, column=18).fill = header_fill
    ws.cell(row=sum_row, column=18).border = thin_border
    ws.cell(row=sum_row, column=18).alignment = center
    
    ws.cell(row=sum_row, column=19, value=f"{total_travel:.0f}").font = bold_font
    ws.cell(row=sum_row, column=19).fill = header_fill
    ws.cell(row=sum_row, column=19).border = thin_border
    ws.cell(row=sum_row, column=19).alignment = center
    
    # 備考欄
    notes_row = sum_row + 2
    ws.merge_cells(f'B{notes_row}:H{notes_row+7}')
    ws.cell(row=notes_row, column=2, value="備考").font = header_font
    ws.cell(row=notes_row, column=2).alignment = top
    ws.cell(row=notes_row, column=2).fill = white_fill
    ws.cell(row=notes_row, column=2).border = thin_border
    
    # 備考欄の罫線
    for r in range(notes_row, notes_row+8):
        for c in range(2, 9):
            ws.cell(row=r, column=c).border = thin_border
    
    # 下部情報
    info_row = notes_row
    ws.cell(row=info_row, column=13, value="実働時間合計").font = header_font
    ws.cell(row=info_row, column=13).alignment = left
    ws.cell(row=info_row, column=14, value=f"{total_work/60:.2f} h").font = normal_font
    ws.cell(row=info_row, column=14).alignment = left
    
    ws.cell(row=info_row+2, column=13, value="自").font = normal_font
    ws.cell(row=info_row+2, column=13).alignment = left
    ws.cell(row=info_row+2, column=14, value=f"{year}年{month}月1日").font = normal_font
    ws.cell(row=info_row+2, column=14).alignment = left
    
    ws.cell(row=info_row+3, column=13, value="至").font = normal_font
    ws.cell(row=info_row+3, column=13).alignment = left
    ws.cell(row=info_row+3, column=14, value=f"{year}年{month}月{days[-1].day}日").font = normal_font
    ws.cell(row=info_row+3, column=14).alignment = left
    
    # ファイル保存
    filename = f"作業時間報告書_{year}年{month}月_{user_display_name or '氏名未入力'}.xlsx"
    wb.save(filename)
    return filename

# ルート定義
@app.route('/auth', methods=['GET', 'POST'])
def auth():
    """ログイン・新規登録画面"""
    auth_mgr = get_auth_manager()
    
    if request.method == 'POST':
        action = request.form.get('action')
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        if action == 'login':
            # ログイン処理
            if auth_mgr.verify_password(username, password):
                auth_mgr.login_user(username)
                return redirect(url_for('index'))
            else:
                return render_template('auth.html', error_message='ユーザー名またはパスワードが間違っています')
        
        elif action == 'register':
            # 新規登録処理
            display_name = request.form.get('display_name', '').strip()
            confirm_password = request.form.get('confirm_password', '')
            
            print(f"DEBUG: 新規登録試行 - ユーザー名={username}, 表示名={display_name}")
            print(f"DEBUG: Firestore利用可能={firestore_manager.is_available()}")
            
            # バリデーション
            if not username or not password or not display_name:
                print("ERROR: 必須項目が未入力")
                return render_template('auth.html', error_message='すべての項目を入力してください')
            elif len(password) < 6:
                print("ERROR: パスワードが短すぎる")
                return render_template('auth.html', error_message='パスワードは6文字以上で設定してください')
            elif password != confirm_password:
                print("ERROR: パスワード確認が一致しない")
                return render_template('auth.html', error_message='パスワードが一致しません')
            
            # 詳細なデバッグ情報を取得
            try:
                from auth_firestore import firestore_auth_manager
                print(f"DEBUG: 認証キャッシュ内ユーザー数={len(firestore_auth_manager.users_cache)}")
                print(f"DEBUG: 認証キャッシュ内ユーザー一覧={list(firestore_auth_manager.users_cache.keys())}")
                
                # Firestoreから直接確認
                if firestore_manager.is_available():
                    existing_user = firestore_manager.get_document('users', username)
                    print(f"DEBUG: Firestore直接確認結果={existing_user is not None}")
                    if existing_user:
                        print(f"DEBUG: 既存ユーザー情報={existing_user}")
                        
            except Exception as debug_e:
                print(f"DEBUG: デバッグ情報取得エラー={str(debug_e)}")
            
            # 新規ユーザー登録
            print(f"DEBUG: ユーザー追加処理開始 - {username}")
            if firestore_auth_manager.add_user(username, password, display_name):
                print(f"DEBUG: ユーザー追加成功 - {username}")
                return render_template('auth.html', 
                                     success_message=f'ユーザー「{display_name}」を登録しました。ログインしてください。')
            else:
                print(f"DEBUG: ユーザー追加失敗 - {username}")
                error_message = f'ユーザー名「{username}」は既に存在します。別のユーザー名を選択してください。'
                
                # 詳細なエラー情報を追加
                try:
                    if firestore_manager.is_available():
                        existing_user = firestore_manager.get_document('users', username)
                        if existing_user:
                            existing_display_name = existing_user.get('display_name', 'unknown')
                            error_message += f' (既存の表示名: {existing_display_name})'
                except Exception as e:
                    error_message += f' (エラー詳細確認失敗: {str(e)})'
                
                return render_template('auth.html', error_message=error_message)
    
    # 既にログイン済みの場合はホームにリダイレクト
    if auth_mgr.is_logged_in():
        return redirect(url_for('index'))
    
    return render_template('auth.html')

@app.route('/logout')
def logout():
    """ログアウト"""
    auth_mgr = get_auth_manager()
    auth_mgr.logout_user()
    return redirect(url_for('auth'))

@app.route('/')
@login_required_decorator
def index():
    """勤怠打刻画面（ホーム）"""
    from datetime import timezone
    
    auth_mgr = get_auth_manager()
    
    # 日本時間で今日の日付を取得
    jst = timezone(timedelta(hours=9))
    today = datetime.now(jst).strftime('%Y-%m-%d')
    
    # 現在のユーザーのデータを取得
    current_user = auth_mgr.get_current_user()
    data = load_user_data(current_user)
    today_data = data.get(today, {})
    
    return render_template('punch.html', 
                         today=today,
                         current_user=current_user,
                         current_display_name=auth_mgr.get_current_display_name(),
                         check_in=today_data.get('check_in', ''),
                         check_out=today_data.get('check_out', ''))

@app.route('/attendance_info')
@login_required_decorator
def attendance_info():
    """勤怠情報ページ"""
    auth_mgr = get_auth_manager()
    attendance_mgr = get_attendance_manager()
    
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))
    
    # 現在のユーザー情報を取得
    current_user = auth_mgr.get_current_user()
    display_name = auth_mgr.get_current_display_name()
    
    # 月の日付範囲を取得
    start_date, end_date = get_month_range(int(year), int(month))
    
    # ユーザー別勤怠データを読み込み
    if attendance_mgr:
        user_data = attendance_mgr.get_user_monthly_data(current_user, year, month)
    else:
        user_data = load_user_data(current_user)
    
    # 日付ごとのデータを準備
    attendance_data = []
    current_date = start_date
    while current_date <= end_date:
        date_str = current_date.strftime('%Y-%m-%d')
        weekday = current_date.strftime('%A')
        is_holiday = check_holiday(current_date)
        
        attendance_data.append({
            'date': date_str,
            'display_date': current_date.day,
            'weekday': weekday,
            'is_holiday': is_holiday,
            'data': user_data.get(date_str, {})
        })
        current_date += timedelta(days=1)
    
    return render_template('attendance_info.html', 
                         attendance_data=attendance_data,
                         year=year, 
                         month=month, 
                         current_user=current_user,
                         display_name=display_name,
                         today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/attendance')
@login_required_decorator
def attendance():
    """勤怠入力ページ"""
    auth_mgr = get_auth_manager()
    attendance_mgr = get_attendance_manager()
    
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))
    
    # 現在のユーザー情報を取得
    current_user = auth_mgr.get_current_user()
    display_name = auth_mgr.get_current_display_name()
    
    # 月の日付範囲を取得
    start_date, end_date = get_month_range(int(year), int(month))
    
    # ユーザー別勤怠データを読み込み
    if attendance_mgr:
        user_data = attendance_mgr.get_user_monthly_data(current_user, year, month)
    else:
        user_data = load_user_data(current_user)
    
    # 日付ごとのデータを準備
    attendance_data = []
    current_date = start_date
    while current_date <= end_date:
        date_str = current_date.strftime('%Y-%m-%d')
        weekday = current_date.strftime('%A')
        is_holiday = check_holiday(current_date)
        
        attendance_data.append({
            'date': date_str,
            'display_date': current_date.day,
            'weekday': weekday,
            'is_holiday': is_holiday,
            'data': user_data.get(date_str, {})
        })
        current_date += timedelta(days=1)
    
    return render_template('attendance.html', 
                         attendance_data=attendance_data,
                         year=year, 
                         month=month, 
                         current_user=current_user,
                         display_name=display_name,
                         today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/save_attendance', methods=['POST'])
@login_required_decorator
def save_attendance():
    """勤怠データを保存（フォーム送信）"""
    auth_mgr = get_auth_manager()
    attendance_mgr = get_attendance_manager()
    current_user = auth_mgr.get_current_user()
    
    if attendance_mgr:
        # Firestore版
        for key, value in request.form.items():
            # 例: key = 'check_in_2025-07-08'
            if '_' in key and len(key.split('_')) >= 2:
                parts = key.split('_')
                field = parts[0]
                date_str = '_'.join(parts[1:])  # 日付に_が含まれる場合に対応
                attendance_mgr.update_user_attendance_data(current_user, date_str, field, value)
    else:
        # 従来版
        data = load_data()
        for key, value in request.form.items():
            if '_' in key:
                field, date_str = key.split('_', 1)
                if date_str not in data:
                    data[date_str] = {}
                data[date_str][field] = value
        save_data(data)
    
    return redirect(url_for('attendance'))

@app.route('/export_excel')
@login_required_decorator
def export_excel():
    """Excelファイルをエクスポート"""
    auth_mgr = get_auth_manager()
    
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))
    
    # ログインユーザーの情報を取得
    current_user = auth_mgr.get_current_user()
    display_name = auth_mgr.get_current_display_name()
    
    # ユーザー別データを取得
    user_data = load_user_data(current_user)
    filename = create_excel_report(year, month, user_data, display_name)
    
    return send_file(filename, as_attachment=True, download_name=filename)

@app.route('/api/save_attendance', methods=['POST'])
@login_required_decorator
def api_save_attendance():
    """API経由で勤怠データを保存"""
    auth_mgr = get_auth_manager()
    attendance_mgr = get_attendance_manager()
    current_user = auth_mgr.get_current_user()
    
    request_data = request.get_json()
    date_str = request_data.get('date')
    field = request_data.get('field')
    value = request_data.get('value')
    
    if attendance_mgr:
        # Firestore版
        success = attendance_mgr.update_user_attendance_data(current_user, date_str, field, value)
        return jsonify({'success': success})
    else:
        # 従来版
        data = load_data()
        if date_str not in data:
            data[date_str] = {}
        data[date_str][field] = value
        save_data(data)
        return jsonify({'success': True})

@app.route('/api/punch', methods=['POST'])
@login_required_decorator
def api_punch():
    """打刻API"""
    try:
        print("DEBUG: 打刻API呼び出し開始")
        auth_mgr = get_auth_manager()
        attendance_mgr = get_attendance_manager()
        
        current_user = auth_mgr.get_current_user()
        print(f"DEBUG: 現在のユーザー={current_user}")
        print(f"DEBUG: リクエストメソッド={request.method}")
        print(f"DEBUG: Content-Type={request.content_type}")
        
        req = request.get_json()
        print(f"DEBUG: リクエストJSON={req}")
        
        if not req:
            print("ERROR: JSONデータが空です")
            return jsonify({'success': False, 'error': 'No JSON data'}), 400
            
        date_str = req.get('date')
        field = req.get('field')  # 'check_in' or 'check_out'
        print(f"DEBUG: date_str={date_str}, field={field}")

        if not date_str or not field:
            print(f"ERROR: 必須パラメータ不足 - date_str={date_str}, field={field}")
            return jsonify({'success': False, 'error': 'Missing parameters'}), 400

        # 現在時刻を日本時間（JST）で取得し、15分単位で四捨五入
        from datetime import timezone
        
        # 日本時間（UTC+9）を取得
        jst = timezone(timedelta(hours=9))
        now = datetime.now(jst)
        print(f"DEBUG: 現在時刻（JST）={now.strftime('%Y-%m-%d %H:%M:%S')}")
        
        minute = now.minute
        # 四捨五入: 0-7→0, 8-22→15, 23-37→30, 38-52→45, 53-59→+1h,0
        round_min = int(15 * round(minute / 15))
        if round_min == 60:
            punch_time = now.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
        else:
            punch_time = now.replace(minute=round_min, second=0, microsecond=0)
        time_str = punch_time.strftime('%H:%M')
        print(f"DEBUG: 打刻時刻（JST）={time_str}")

        # データ保存
        if attendance_mgr:
            # Firestore版
            success = attendance_mgr.update_user_attendance_data(current_user, date_str, field, time_str)
            if success:
                print("DEBUG: Firestore勤怠データ保存完了")
            else:
                print("ERROR: Firestore勤怠データ保存失敗")
        else:
            # 従来版
            user_data = load_user_data(current_user)
            print("DEBUG: ユーザーデータロード完了")
            
            if date_str not in user_data:
                user_data[date_str] = {}
            user_data[date_str][field] = time_str
            print(f"DEBUG: 保存前ユーザーデータ={user_data.get(date_str, {})}")
            
            save_user_data(current_user, user_data)
            print("DEBUG: ユーザーデータ保存完了")
        
        return jsonify({'success': True, 'time': time_str})
        
    except Exception as e:
        print(f"ERROR: 打刻API例外発生 - {str(e)}")
        import traceback
        print(f"ERROR: トレースバック - {traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/save_field', methods=['POST'])
@login_required_decorator
def api_save_field():
    """フィールド保存API"""
    auth_mgr = get_auth_manager()
    attendance_mgr = get_attendance_manager()
    current_user = auth_mgr.get_current_user()
    
    req = request.get_json()
    date_str = req.get('date')
    field = req.get('field')
    value = req.get('value')
    
    if attendance_mgr:
        # Firestore版
        success = attendance_mgr.update_user_attendance_data(current_user, date_str, field, value)
        return jsonify({'success': success})
    else:
        # 従来版
        data = load_data()
        if date_str not in data:
            data[date_str] = {}
        data[date_str][field] = value
        save_data(data)
        return jsonify({'success': True})

# データ移行エンドポイント（管理者用）
@app.route('/admin/migrate_to_firestore', methods=['POST'])
@login_required_decorator
def migrate_to_firestore():
    """従来のデータをFirestoreに移行"""
    auth_mgr = get_auth_manager()
    attendance_mgr = get_attendance_manager()
    
    # 管理者権限チェック（簡単な実装）
    current_user = auth_mgr.get_current_user()
    if current_user != 'admin':  # 実際の管理者ユーザー名に変更
        return jsonify({'success': False, 'error': 'Admin access required'}), 403
    
    if not attendance_mgr:
        return jsonify({'success': False, 'error': 'Firestore not available'}), 500
    
    try:
        # 従来形式のデータを読み込み
        legacy_data = load_data()
        
        # Firestoreに移行
        success = attendance_mgr.migrate_from_legacy_format(legacy_data)
        
        return jsonify({'success': success, 'message': 'Migration completed'})
        
    except Exception as e:
        print(f"ERROR: データ移行失敗 - {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/debug/firestore', methods=['GET'])
def api_debug_firestore():
    """Firestore状態をデバッグ用に表示（認証不要）"""
    try:
        debug_info = {
            'timestamp': datetime.now().isoformat(),
            'use_firestore': USE_FIRESTORE,
            'firestore_available': firestore_manager.is_available() if USE_FIRESTORE else False,
            'firebase_project_id': os.environ.get('FIREBASE_PROJECT_ID', 'Not set'),
            'has_credentials': 'GOOGLE_APPLICATION_CREDENTIALS_BASE64' in os.environ,
            'environment': 'vercel' if 'VERCEL' in os.environ else 'local',
            'python_version': sys.version.split()[0],
        }
        
        if USE_FIRESTORE and firestore_manager.is_available():
            try:
                # ユーザー数を取得
                users_docs = firestore_manager.get_collection('users')
                user_count = len(users_docs) if users_docs else 0
                
                # ユーザー一覧を取得
                user_list = []
                if users_docs:
                    for doc in users_docs:
                        user_list.append({
                            'username': doc.get('username', 'unknown'),
                            'display_name': doc.get('display_name', 'unknown'),
                            'created_at': doc.get('created_at', 'unknown')
                        })
                
                # セッション数を取得
                sessions_docs = firestore_manager.get_collection('user_sessions')
                session_count = len(sessions_docs) if sessions_docs else 0
                
                # 勤怠データ数を取得
                attendance_docs = firestore_manager.get_collection('user_attendance')
                attendance_count = len(attendance_docs) if attendance_docs else 0
                
                # 認証管理の情報を取得
                try:
                    from auth_firestore import firestore_auth_manager
                    auth_cache_size = len(firestore_auth_manager.users_cache)
                    auth_cache_users = list(firestore_auth_manager.users_cache.keys())
                except Exception as auth_e:
                    auth_cache_size = 'Error'
                    auth_cache_users = f'Error: {str(auth_e)}'
                
                debug_info.update({
                    'user_count': user_count,
                    'user_list': user_list,
                    'session_count': session_count,
                    'attendance_count': attendance_count,
                    'auth_cache_size': auth_cache_size,
                    'auth_cache_users': auth_cache_users
                })
                
            except Exception as e:
                debug_info['firestore_error'] = str(e)
        
        return jsonify(debug_info)
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/debug/firestore', methods=['GET'])
def debug_firestore():
    """Firestore状態をデバッグ用に表示"""
    if not USE_FIRESTORE:
        return jsonify({"error": "Firestore is disabled"}), 400
    
    try:
        # ユーザー数を取得
        users_docs = firestore_manager.get_collection('users')
        user_count = len(users_docs) if users_docs else 0
        
        # ユーザー一覧を取得
        user_list = []
        if users_docs:
            for doc in users_docs:
                user_list.append({
                    'username': doc.get('username', 'unknown'),
                    'display_name': doc.get('display_name', 'unknown'),
                    'created_at': doc.get('created_at', 'unknown')
                })
        
        # セッション数を取得
        sessions_docs = firestore_manager.get_collection('user_sessions')
        session_count = len(sessions_docs) if sessions_docs else 0
        
        # 勤怠データ数を取得
        attendance_docs = firestore_manager.get_collection('user_attendance')
        attendance_count = len(attendance_docs) if attendance_docs else 0
        
        debug_info = {
            'firestore_available': firestore_manager.is_available(),
            'use_firestore': USE_FIRESTORE,
            'firebase_project_id': os.environ.get('FIREBASE_PROJECT_ID', 'Not set'),
            'has_credentials': 'GOOGLE_APPLICATION_CREDENTIALS_BASE64' in os.environ,
            'user_count': user_count,
            'user_list': user_list,
            'session_count': session_count,
            'attendance_count': attendance_count,
            'auth_cache_size': len(auth_manager.users_cache),
            'auth_cache_users': list(auth_manager.users_cache.keys())
        }
        
        return jsonify(debug_info)
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    print(f"DEBUG: Firestore利用可能 = {firestore_manager.is_available()}")
    print(f"DEBUG: USE_FIRESTORE = {USE_FIRESTORE}")
    
    # 開発時はデバッグモードで実行
    app.run(debug=True, host='0.0.0.0', port=5001) 