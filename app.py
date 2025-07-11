from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify, session
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import json
from dateutil.relativedelta import relativedelta
from math import floor
import requests
from firebase_config import firebase_db
from auth import auth_manager, login_required

# jpholidayのインポートを安全に行う
try:
    import jpholiday
    JPHOLIDAY_AVAILABLE = True
except (ImportError, TypeError) as e:
    print(f"jpholidayインポートエラー: {e}")
    JPHOLIDAY_AVAILABLE = False

app = Flask(__name__)
# セッション用の秘密鍵（環境変数から取得、なければランダム生成）
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# データファイルのパス
DATA_FILE = 'attendance_data.json'

# Vercel環境でのデータ永続化のための設定
GIST_ID = os.environ.get('GIST_ID')
GITHUB_TOKEN = os.environ.get('GITHUB_TOKEN')
USE_GIST = bool(GIST_ID and GITHUB_TOKEN)

# 全テンプレートで現在の年を利用できるようにする
@app.context_processor
def inject_now():
    from datetime import datetime
    return {'now': datetime.now()}

def load_user_data(username: str):
    """特定ユーザーの勤怠データを読み込む"""
    all_data = load_data()
    return all_data.get('users', {}).get(username, {})

def save_user_data(username: str, user_data: dict):
    """特定ユーザーの勤怠データを保存する"""
    all_data = load_data()
    if 'users' not in all_data:
        all_data['users'] = {}
    if username not in all_data['users']:
        all_data['users'][username] = {}
    
    all_data['users'][username] = user_data
    save_data(all_data)

def load_data():
    """全体勤怠データを読み込む"""
    print(f"DEBUG: Firebase利用可能 = {firebase_db.is_available()}")
    print(f"DEBUG: GIST_ID = {GIST_ID}")
    print(f"DEBUG: GITHUB_TOKEN = {'設定済み' if GITHUB_TOKEN else '未設定'}")
    print(f"DEBUG: USE_GIST = {USE_GIST}")
    
    # Firebase優先で試行
    if firebase_db.is_available():
        data = firebase_db.load_data()
        if data:  # Firebaseから正常にデータを取得できた場合
            return data
    
    # Firebaseが失敗した場合、Gistを試行
    if USE_GIST:
        return load_data_from_gist()
    
    # 最後にローカルファイル（開発環境のみ）
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    return {}

def save_data(data):
    """全体勤怠データを保存する"""
    saved = False
    
    # Firebase優先で保存
    if firebase_db.is_available():
        if firebase_db.save_data(data):
            saved = True
            print("DEBUG: Firebaseに保存成功")
    
    # Firebaseが失敗した場合、Gistに保存
    if not saved and USE_GIST:
        save_data_to_gist(data)
        saved = True
        print("DEBUG: Gistに保存成功")
    
    # 開発環境ではローカルファイルにも保存
    if not saved:
        try:
            with open(DATA_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print("DEBUG: ローカルファイルに保存成功")
        except Exception as e:
            print(f"ERROR: ローカル保存失敗 - {str(e)}")

def load_data_from_gist():
    """GitHub Gistからデータを読み込む"""
    try:
        headers = {
            'Authorization': f'token {GITHUB_TOKEN}',
            'Accept': 'application/vnd.github.v3+json'
        }
        response = requests.get(f'https://api.github.com/gists/{GIST_ID}', headers=headers)
        if response.status_code == 200:
            gist_data = response.json()
            files = gist_data.get('files', {})
            for filename, file_data in files.items():
                if filename == 'attendance_data.json':
                    content = file_data.get('content', '{}')
                    return json.loads(content)
        return {}
    except Exception as e:
        print(f"Gist読み込みエラー: {e}")
        return {}

def save_data_to_gist(data):
    """GitHub Gistにデータを保存する"""
    try:
        print(f"DEBUG: Gist保存開始 - データサイズ: {len(json.dumps(data))}")
        headers = {
            'Authorization': f'token {GITHUB_TOKEN}',
            'Accept': 'application/vnd.github.v3+json'
        }
        
        # 現在のGistを取得
        response = requests.get(f'https://api.github.com/gists/{GIST_ID}', headers=headers)
        print(f"DEBUG: Gist取得レスポンス: {response.status_code}")
        if response.status_code != 200:
            print(f"Gist取得エラー: {response.status_code}")
            print(f"DEBUG: レスポンス内容: {response.text}")
            return
        
        gist_data = response.json()
        files = gist_data.get('files', {})
        
        # ファイル内容を更新
        files['attendance_data.json'] = {
            'content': json.dumps(data, ensure_ascii=False, indent=2)
        }
        
        # Gistを更新
        update_data = {
            'files': files
        }
        
        response = requests.patch(
            f'https://api.github.com/gists/{GIST_ID}',
            headers=headers,
            json=update_data
        )
        
        print(f"DEBUG: Gist更新レスポンス: {response.status_code}")
        if response.status_code != 200:
            print(f"Gist更新エラー: {response.status_code}")
            print(f"DEBUG: レスポンス内容: {response.text}")
        else:
            print("DEBUG: Gist更新成功")
            
    except Exception as e:
        print(f"Gist保存エラー: {e}")
        import traceback
        print(f"DEBUG: エラー詳細: {traceback.format_exc()}")

def check_holiday(date):
    """祝日判定（jpholidayが利用できない場合は土日のみ判定）"""
    if JPHOLIDAY_AVAILABLE:
        try:
            return jpholiday.is_holiday(date)
        except:
            pass
    # jpholidayが利用できない場合は土日のみ判定
    return date.weekday() >= 5  # 土曜日(5)と日曜日(6)

def get_month_range(year, month):
    """指定された年月の日付範囲を取得"""
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = datetime(year, month + 1, 1) - timedelta(days=1)
    return start_date, end_date

def create_excel_report(year, month, data, username):
    import calendar
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.dimensions import RowDimension
    
    def to_wareki(y, m):
        if y > 2019 or (y == 2019 and m >= 5):
            return f"令和 {y-2018}"
        else:
            return f"{y}年"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "template"
    
    # スタイル
    title_font = Font(name='MS Gothic', size=16, bold=True)
    header_font = Font(name='MS Gothic', size=10, bold=True)
    normal_font = Font(name='MS Gothic', size=9)
    red_font = Font(name='MS Gothic', size=9, color="FF0000", bold=True)
    bold_font = Font(name='MS Gothic', size=9, bold=True)
    
    # 色
    table_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    header_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
    company_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # 罫線
    thin = Side(style='thin', color='000000')
    thick = Side(style='medium', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    thick_border = Border(left=thick, right=thick, top=thick, bottom=thick)
    
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    top = Alignment(horizontal='left', vertical='top')
    
    # 列幅・行高（JとKの間に空白列を追加、交通費合計用）
    # [A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T]
    #  A:未使用, B:日付, ... J:備考, K:空白, L:日付(右), ... S:備考(右), T:空白
    col_widths = [5, 10, 12, 14, 14, 14, 10, 14, 14, 18, 4, 10, 12, 14, 14, 14, 10, 14, 14, 18, 4]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w
    for r in range(1, 60):
        ws.row_dimensions[r].height = 22
    ws.row_dimensions[2].height = 32
    
    # タイトル（広く結合）
    ws.merge_cells('B2:P2')
    ws['B2'] = "作業時間報告書"
    ws['B2'].font = title_font
    ws['B2'].alignment = center
    
    # 年月度（merge_cellsを使わず個別セルに値を書き込む）
    # ws.merge_cells('B4:D4') などは削除済み
    ws['B4'] = to_wareki(year, month)
    ws['C4'] = f"{year}"
    ws['D4'] = "年"
    ws['E4'] = f"{month}"
    ws['F4'] = "月度"
    ws['B4'].font = header_font
    ws['C4'].font = header_font
    ws['D4'].font = header_font
    ws['E4'].font = header_font
    ws['F4'].font = header_font
    ws['B4'].alignment = center
    ws['C4'].alignment = center
    ws['D4'].alignment = center
    ws['E4'].alignment = center
    ws['F4'].alignment = center

    # 固定情報（和暦・年・月度・対応客先名・会社名・氏名など）をコンパクトに
    # 年月度
    ws['B4'] = to_wareki(year, month)
    ws['C4'] = f"{year}"
    ws['D4'] = "年"
    ws['E4'] = f"{month}"
    ws['F4'] = "月度"
    ws['B4'].font = header_font
    ws['C4'].font = header_font
    ws['D4'].font = header_font
    ws['E4'].font = header_font
    ws['F4'].font = header_font
    ws['B4'].alignment = center
    ws['C4'].alignment = center
    ws['D4'].alignment = center
    ws['E4'].alignment = center
    ws['F4'].alignment = center
    # 固定情報
    ws['B7'] = "対応客先名"
    ws['C7'] = "株式会社LINE"
    ws['B7'].font = header_font
    ws['C7'].font = normal_font
    ws['B7'].alignment = center
    ws['C7'].alignment = left
    ws['C7'].fill = company_fill
    ws['F7'] = "会社名"
    ws['G7'] = "HIGHFLAT"
    ws['F7'].font = header_font
    ws['G7'].font = normal_font
    ws['F7'].alignment = center
    ws['G7'].alignment = left
    ws['G7'].fill = company_fill
    ws['I7'] = "氏名"
    ws['J7'] = username or ""
    ws['I7'].font = header_font
    ws['J7'].font = normal_font
    ws['I7'].alignment = center
    ws['J7'].alignment = left
    ws['J7'].fill = company_fill
    
    # テーブルヘッダー
    headers = ["日付", "曜日", "出勤時間", "退勤時間", "実働時間", "交通費", "出発駅", "目的駅", "備考"]
    for i, h in enumerate(headers):
        ws.cell(row=13, column=2+i, value=h).font = header_font
        ws.cell(row=13, column=2+i).alignment = center
        ws.cell(row=13, column=2+i).fill = header_fill
        ws.cell(row=13, column=2+i).border = thin_border
        ws.cell(row=13, column=12+i, value=h).font = header_font
        ws.cell(row=13, column=12+i).alignment = center
        ws.cell(row=13, column=12+i).fill = header_fill
        ws.cell(row=13, column=12+i).border = thin_border
    
    # 日付データ
    start_date, end_date = get_month_range(year, month)
    days = []
    current_date = start_date
    while current_date <= end_date:
        days.append(current_date)
        current_date += timedelta(days=1)
    left_days = days[:16]
    right_days = days[16:]
    total_work = 0
    total_travel = 0
    # 左側
    for idx, d in enumerate(left_days):
        r = 14 + idx
        date_str = d.strftime('%Y-%m-%d')
        weekday = d.strftime('%a')
        jp_week = {'Mon':'月','Tue':'火','Wed':'水','Thu':'木','Fri':'金','Sat':'土','Sun':'日'}[weekday]
        att = data.get(date_str, {})
        check_in = att.get('check_in', '')
        check_out = att.get('check_out', '')
        break_time = float(att.get('break_time', '1.0') or 1.0)
        notes = att.get('notes', '')
        travel_cost = att.get('travel_cost', '')
        travel_from = att.get('travel_from', '')
        travel_to = att.get('travel_to', '')
        is_holiday = check_holiday(d)
        def get_minutes(t):
            if not t: return None
            h, m = map(int, t.split(':'))
            return h*60 + m
        work_min = None
        if check_in and check_out:
            work_min = get_minutes(check_out) - get_minutes(check_in) - int(break_time*60)
            if work_min < 0: work_min = None
        total_work += work_min or 0
        try:
            total_travel += float(travel_cost) if travel_cost else 0
        except:
            pass
        ws.cell(row=r, column=2, value=d.day).font = normal_font
        # 祝日・土日判定
        if jp_week in ['土','日'] or is_holiday:
            ws.cell(row=r, column=3, value=jp_week).font = red_font
            ws.cell(row=r, column=3).fill = table_fill
        else:
            ws.cell(row=r, column=3, value=jp_week).font = normal_font
            ws.cell(row=r, column=3).fill = table_fill
        ws.cell(row=r, column=4, value=check_in).font = normal_font
        ws.cell(row=r, column=5, value=check_out).font = normal_font
        ws.cell(row=r, column=6, value=(f"{work_min/60:.2f}" if work_min is not None else "")).font = normal_font
        ws.cell(row=r, column=7, value=travel_cost).font = normal_font
        ws.cell(row=r, column=8, value=travel_from).font = normal_font
        ws.cell(row=r, column=9, value=travel_to).font = normal_font
        ws.cell(row=r, column=10, value=notes).font = normal_font
        for i in range(9):
            c = ws.cell(row=r, column=2+i)
            c.fill = table_fill
            c.border = thin_border
            c.alignment = center if i not in [9] else left
        ws.cell(row=r, column=11).fill = white_fill  # 空白列
    # 右側
    for idx, d in enumerate(right_days):
        r = 14 + idx
        date_str = d.strftime('%Y-%m-%d')
        weekday = d.strftime('%a')
        jp_week = {'Mon':'月','Tue':'火','Wed':'水','Thu':'木','Fri':'金','Sat':'土','Sun':'日'}[weekday]
        att = data.get(date_str, {})
        check_in = att.get('check_in', '')
        check_out = att.get('check_out', '')
        break_time = float(att.get('break_time', '1.0') or 1.0)
        notes = att.get('notes', '')
        travel_cost = att.get('travel_cost', '')
        travel_from = att.get('travel_from', '')
        travel_to = att.get('travel_to', '')
        is_holiday = check_holiday(d)
        def get_minutes(t):
            if not t: return None
            h, m = map(int, t.split(':'))
            return h*60 + m
        work_min = None
        if check_in and check_out:
            work_min = get_minutes(check_out) - get_minutes(check_in) - int(break_time*60)
            if work_min < 0: work_min = None
        total_work += work_min or 0
        try:
            total_travel += float(travel_cost) if travel_cost else 0
        except:
            pass
        ws.cell(row=r, column=12, value=d.day).font = normal_font
        # 祝日・土日判定
        if jp_week in ['土','日'] or is_holiday:
            ws.cell(row=r, column=13, value=jp_week).font = red_font
            ws.cell(row=r, column=13).fill = table_fill
        else:
            ws.cell(row=r, column=13, value=jp_week).font = normal_font
            ws.cell(row=r, column=13).fill = table_fill
        ws.cell(row=r, column=14, value=check_in).font = normal_font
        ws.cell(row=r, column=15, value=check_out).font = normal_font
        ws.cell(row=r, column=16, value=(f"{work_min/60:.2f}" if work_min is not None else "")).font = normal_font
        ws.cell(row=r, column=17, value=travel_cost).font = normal_font
        ws.cell(row=r, column=18, value=travel_from).font = normal_font
        ws.cell(row=r, column=19, value=travel_to).font = normal_font
        ws.cell(row=r, column=20, value=notes).font = normal_font
        for i in range(9):
            c = ws.cell(row=r, column=12+i)
            c.fill = table_fill
            c.border = thin_border
            c.alignment = center if i not in [9] else left
        ws.cell(row=r, column=21).fill = white_fill  # 空白列
    # 合計行（右側のみ）
    sum_row = 14+len(right_days)
    # 実働時間合計（O列・P列）
    ws.cell(row=sum_row, column=15, value="計").font = bold_font
    ws.cell(row=sum_row, column=15).fill = header_fill
    ws.cell(row=sum_row, column=15).border = thin_border
    ws.cell(row=sum_row, column=15).alignment = center
    ws.cell(row=sum_row, column=16, value=f"{total_work/60:.2f}").font = bold_font
    ws.cell(row=sum_row, column=16).fill = header_fill
    ws.cell(row=sum_row, column=16).border = thin_border
    ws.cell(row=sum_row, column=16).alignment = center
    # 交通費合計（R列・S列）
    ws.cell(row=sum_row, column=18, value="交通費合計").font = bold_font
    ws.cell(row=sum_row, column=18).fill = header_fill
    ws.cell(row=sum_row, column=18).border = thin_border
    ws.cell(row=sum_row, column=18).alignment = center
    ws.cell(row=sum_row, column=19, value=f"{total_travel:.2f}").font = bold_font
    ws.cell(row=sum_row, column=19).fill = header_fill
    ws.cell(row=sum_row, column=19).border = thin_border
    ws.cell(row=sum_row, column=19).alignment = center
    # 表の枠線をすべて細線に統一
    for row in range(13, sum_row+1):
        for col in range(2, 21):
            ws.cell(row=row, column=col).border = thin_border
    # 備考欄（広く結合）
    notes_row = max(14+len(left_days), 14+len(right_days)) + 2
    ws.merge_cells(f'B{notes_row}:H{notes_row+7}')
    ws.cell(row=notes_row, column=2, value="備考").font = header_font
    ws.cell(row=notes_row, column=2).alignment = top
    ws.cell(row=notes_row, column=2).fill = white_fill
    ws.cell(row=notes_row, column=2).border = thin_border
    for r in range(notes_row, notes_row+8):
        for c in range(2, 9):
            ws.cell(row=r, column=c).border = thin_border
    # 下部情報（広く配置）
    info_row = notes_row
    ws.cell(row=info_row, column=13, value="受働時間合計").font = header_font
    ws.cell(row=info_row, column=13).alignment = left
    ws.cell(row=info_row, column=14, value=f"{total_work/60:.2f} h").font = normal_font
    ws.cell(row=info_row, column=14).alignment = left
    ws.cell(row=info_row+2, column=13, value="自").font = normal_font
    ws.cell(row=info_row+2, column=13).alignment = left
    ws.cell(row=info_row+2, column=14, value=f"{year}年{month}月1日").font = normal_font
    ws.cell(row=info_row+2, column=14).alignment = left
    ws.cell(row=info_row+3, column=13, value="至").font = normal_font
    ws.cell(row=info_row+3, column=13).alignment = left
    ws.cell(row=info_row+3, column=14, value=f"{year}年{month}月{days[-1].day}日").font = normal_font
    ws.cell(row=info_row+3, column=14).alignment = left
    # 外枠太線（→細線に変更）
    for row in range(13, 14+len(left_days)+1):
        for col in range(2, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
    for row in range(13, 14+len(right_days)+1):
        for col in range(9, 15):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
    filename = f"作業時間報告書_{year}年{month}月_{username or '氏名未入力'}.xlsx"
    wb.save(filename)
    return filename

@app.route('/auth', methods=['GET', 'POST'])
def auth():
    """ログイン・新規登録画面"""
    if request.method == 'POST':
        action = request.form.get('action')
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        if action == 'login':
            # ログイン処理
            if auth_manager.verify_password(username, password):
                auth_manager.login_user(username)
                return redirect(url_for('index'))
            else:
                return render_template('auth.html', error_message='ユーザー名またはパスワードが間違っています')
        
        elif action == 'register':
            # 新規登録処理
            display_name = request.form.get('display_name', '').strip()
            confirm_password = request.form.get('confirm_password', '')
            
            # バリデーション
            if not username or not password or not display_name:
                return render_template('auth.html', error_message='すべての項目を入力してください')
            elif len(password) < 6:
                return render_template('auth.html', error_message='パスワードは6文字以上で設定してください')
            elif password != confirm_password:
                return render_template('auth.html', error_message='パスワードが一致しません')
            elif not username.replace('_', '').isalnum():
                return render_template('auth.html', error_message='ユーザー名は英数字とアンダースコア（_）のみ使用できます')
            else:
                # ユーザー追加
                if auth_manager.add_user(username, password, display_name):
                    auth_manager.login_user(username)
                    return redirect(url_for('index'))
                else:
                    return render_template('auth.html', error_message=f'ユーザー「{username}」は既に存在します')
    
    # 既にログイン済みの場合はホームにリダイレクト
    if auth_manager.is_logged_in():
        return redirect(url_for('index'))
    
    return render_template('auth.html')

@app.route('/logout')
def logout():
    """ログアウト"""
    auth_manager.logout_user()
    return redirect(url_for('auth'))



@app.route('/')
@login_required
def index():
    """勤怠打刻画面（ホーム）"""
    from datetime import timezone
    
    # 日本時間で今日の日付を取得
    jst = timezone(timedelta(hours=9))
    today = datetime.now(jst).strftime('%Y-%m-%d')
    
    # 現在のユーザーのデータを取得
    current_user = auth_manager.get_current_user()
    data = load_user_data(current_user)
    today_data = data.get(today, {})
    
    return render_template('punch.html', 
                         today=today,
                         current_user=current_user,
                         current_display_name=auth_manager.get_current_display_name(),
                         check_in=today_data.get('check_in', ''),
                         check_out=today_data.get('check_out', ''))

@app.route('/attendance_info')
@login_required
def attendance_info():
    """勤怠情報ページ"""
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))
    
    # 現在のユーザー情報を取得
    current_user = auth_manager.get_current_user()
    display_name = auth_manager.get_current_display_name()
    
    # 月の日付範囲を取得
    start_date, end_date = get_month_range(int(year), int(month))
    
    # ユーザー別勤怠データを読み込み
    user_data = load_user_data(current_user)
    
    # 日付ごとのデータを準備
    attendance_data = []
    current_date = start_date
    while current_date <= end_date:
        date_str = current_date.strftime('%Y-%m-%d')
        weekday = current_date.strftime('%A')
        is_holiday = check_holiday(current_date)
        
        attendance_data.append({
            'date': date_str,
            'display_date': current_date.day,
            'weekday': weekday,
            'is_holiday': is_holiday,
            'data': user_data.get(date_str, {})
        })
        current_date += timedelta(days=1)
    
    return render_template('attendance_info.html', 
                         attendance_data=attendance_data,
                         year=year, 
                         month=month, 
                         current_user=current_user,
                         display_name=display_name,
                         today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/attendance')
@login_required
def attendance():
    """勤怠入力ページ"""
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))
    
    # 現在のユーザー情報を取得
    current_user = auth_manager.get_current_user()
    display_name = auth_manager.get_current_display_name()
    
    # 月の日付範囲を取得
    start_date, end_date = get_month_range(int(year), int(month))
    
    # ユーザー別勤怠データを読み込み
    user_data = load_user_data(current_user)
    
    # 日付ごとのデータを準備
    attendance_data = []
    current_date = start_date
    while current_date <= end_date:
        date_str = current_date.strftime('%Y-%m-%d')
        weekday = current_date.strftime('%A')
        is_holiday = check_holiday(current_date)
        
        attendance_data.append({
            'date': date_str,
            'display_date': current_date.day,
            'weekday': weekday,
            'is_holiday': is_holiday,
            'data': user_data.get(date_str, {})
        })
        current_date += timedelta(days=1)
    
    return render_template('attendance.html', 
                         attendance_data=attendance_data,
                         year=year, 
                         month=month, 
                         current_user=current_user,
                         display_name=display_name,
                         today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/save_attendance', methods=['POST'])
def save_attendance():
    data = load_data()
    for key, value in request.form.items():
        # 例: key = 'check_in_2025-07-08'
        if '_' in key:
            field, date_str = key.split('_', 1)
            if date_str not in data:
                data[date_str] = {}
            data[date_str][field] = value
    save_data(data)
    return redirect(url_for('attendance'))

@app.route('/export_excel')
@login_required
def export_excel():
    """Excelファイルをエクスポート"""
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))
    
    # ログインユーザーの情報を取得
    current_user = auth_manager.get_current_user()
    display_name = auth_manager.get_current_display_name()
    
    # ユーザー別データを取得
    user_data = load_user_data(current_user)
    filename = create_excel_report(year, month, user_data, display_name)
    
    return send_file(filename, as_attachment=True, download_name=filename)

@app.route('/api/save_attendance', methods=['POST'])
def api_save_attendance():
    """API経由で勤怠データを保存"""
    data = load_data()
    request_data = request.get_json()
    
    date_str = request_data.get('date')
    field = request_data.get('field')
    value = request_data.get('value')
    
    if date_str not in data:
        data[date_str] = {}
    
    data[date_str][field] = value
    save_data(data)
    
    return jsonify({'success': True})

@app.route('/api/punch', methods=['POST'])
@login_required
def api_punch():
    try:
        print("DEBUG: 打刻API呼び出し開始")
        current_user = auth_manager.get_current_user()
        print(f"DEBUG: 現在のユーザー={current_user}")
        print(f"DEBUG: リクエストメソッド={request.method}")
        print(f"DEBUG: Content-Type={request.content_type}")
        
        # 現在のユーザーのデータを取得
        user_data = load_user_data(current_user)
        print("DEBUG: ユーザーデータロード完了")
        
        req = request.get_json()
        print(f"DEBUG: リクエストJSON={req}")
        
        if not req:
            print("ERROR: JSONデータが空です")
            return jsonify({'success': False, 'error': 'No JSON data'}), 400
            
        date_str = req.get('date')
        field = req.get('field')  # 'check_in' or 'check_out'
        print(f"DEBUG: date_str={date_str}, field={field}")

        if not date_str or not field:
            print(f"ERROR: 必須パラメータ不足 - date_str={date_str}, field={field}")
            return jsonify({'success': False, 'error': 'Missing parameters'}), 400

        # 現在時刻を日本時間（JST）で取得し、15分単位で四捨五入
        from datetime import timezone
        
        # 日本時間（UTC+9）を取得
        jst = timezone(timedelta(hours=9))
        now = datetime.now(jst)
        print(f"DEBUG: 現在時刻（JST）={now.strftime('%Y-%m-%d %H:%M:%S')}")
        
        minute = now.minute
        # 四捨五入: 0-7→0, 8-22→15, 23-37→30, 38-52→45, 53-59→+1h,0
        round_min = int(15 * round(minute / 15))
        if round_min == 60:
            punch_time = now.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
        else:
            punch_time = now.replace(minute=round_min, second=0, microsecond=0)
        time_str = punch_time.strftime('%H:%M')
        print(f"DEBUG: 打刻時刻（JST）={time_str}")

        # ユーザー別にデータを保存
        if date_str not in user_data:
            user_data[date_str] = {}
        user_data[date_str][field] = time_str
        print(f"DEBUG: 保存前ユーザーデータ={user_data.get(date_str, {})}")
        
        save_user_data(current_user, user_data)
        print("DEBUG: ユーザーデータ保存完了")
        
        return jsonify({'success': True, 'time': time_str})
        
    except Exception as e:
        print(f"ERROR: 打刻API例外発生 - {str(e)}")
        import traceback
        print(f"ERROR: トレースバック - {traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/save_field', methods=['POST'])
def api_save_field():
    data = load_data()
    req = request.get_json()
    date_str = req.get('date')
    field = req.get('field')
    value = req.get('value')
    if date_str not in data:
        data[date_str] = {}
    data[date_str][field] = value
    save_data(data)
    return jsonify({'success': True})

@app.route('/debug_env')
def debug_env():
    """環境変数デバッグ用"""
    return jsonify({
        'FIREBASE_URL': os.environ.get('FIREBASE_URL'),
        'FIREBASE_AVAILABLE': firebase_db.is_available(),
        'GIST_ID': GIST_ID,
        'GITHUB_TOKEN_SET': bool(GITHUB_TOKEN),
        'GITHUB_TOKEN_PREFIX': GITHUB_TOKEN[:20] + '...' if GITHUB_TOKEN else None,
        'USE_GIST': USE_GIST,
        'environment': os.environ.get('VERCEL', 'local')
    })

@app.route('/test_api')
def test_api():
    """APIテスト用ページ"""
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>API Test</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            .result { margin: 20px 0; padding: 10px; border: 1px solid #ccc; background: #f9f9f9; }
            button { padding: 10px 20px; font-size: 16px; margin: 10px 0; }
            pre { background: #f0f0f0; padding: 10px; overflow-x: auto; }
        </style>
    </head>
    <body>
        <h1>API Test for Vercel</h1>
        <button onclick="testPunchAPI()">Test Punch API</button>
        <button onclick="testLoadData()">Test Load Data</button>
        <div id="result"></div>

        <script>
        async function testPunchAPI() {
            const resultDiv = document.getElementById('result');
            resultDiv.innerHTML = '<p>Testing Punch API...</p>';
            
            try {
                const response = await fetch('/api/punch', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    },
                    body: JSON.stringify({
                        date: '2025-01-10',
                        field: 'check_in'
                    })
                });
                
                const responseText = await response.text();
                console.log('Raw response:', responseText);
                
                resultDiv.innerHTML = `
                    <div class="result">
                        <h3>Punch API Test Result</h3>
                        <p><strong>Status:</strong> ${response.status} ${response.statusText}</p>
                        <p><strong>Response:</strong></p>
                        <pre>${responseText}</pre>
                    </div>
                `;
                
                if (response.ok) {
                    try {
                        const data = JSON.parse(responseText);
                        console.log('Parsed response:', data);
                    } catch (e) {
                        console.error('Failed to parse JSON:', e);
                    }
                }
                
            } catch (error) {
                console.error('Error:', error);
                resultDiv.innerHTML = `
                    <div class="result">
                        <h3>Error</h3>
                        <p>${error.message}</p>
                    </div>
                `;
            }
        }
        
        async function testLoadData() {
            const resultDiv = document.getElementById('result');
            resultDiv.innerHTML = '<p>Testing Load Data...</p>';
            
            try {
                const response = await fetch('/', {
                    method: 'GET'
                });
                
                const responseText = await response.text();
                
                resultDiv.innerHTML = `
                    <div class="result">
                        <h3>Load Data Test Result</h3>
                        <p><strong>Status:</strong> ${response.status} ${response.statusText}</p>
                        <p><strong>Page loaded successfully:</strong> ${responseText.includes('勤怠打刻') ? 'Yes' : 'No'}</p>
                    </div>
                `;
                
            } catch (error) {
                console.error('Error:', error);
                resultDiv.innerHTML = `
                    <div class="result">
                        <h3>Error</h3>
                        <p>${error.message}</p>
                    </div>
                `;
            }
        }
        </script>
    </body>
    </html>
    '''

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001) 